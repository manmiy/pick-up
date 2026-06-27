# =========================================================
# 自動発注・連絡書システム v2.0
# =========================================================
# ワークフロー:
#   Step 1: 拾い出し表Excelアップロード
#   Step 2: カタログ画像閲覧
#   Step 3: 材料の検索 & チェック
#   Step 4: 発注先・担当者・日付の入力
#   Step 5: 発注書・連絡書Excelダウンロード
#   Step 6: ExcelアップロードしてPDF変換
# =========================================================

import streamlit as st
import pandas as pd
import openpyxl
import openpyxl.cell.cell
import os
import io
import platform
import subprocess
import zipfile
import posixpath
from xml.etree import ElementTree as ET
from datetime import datetime, date

# =========================================================
# ページ設定
# =========================================================
st.set_page_config(
    page_title="自動発注システム",
    page_icon="📋",
    layout="wide",
    initial_sidebar_state="expanded",
)

# =========================================================
# カスタムCSS
# =========================================================
st.markdown(
    """
<style>
@import url('https://fonts.googleapis.com/css2?family=Noto+Sans+JP:wght@300;400;500;600;700&display=swap');

html, body, [class*="st-"] {
    font-family: 'Noto Sans JP', sans-serif !important;
}

/* ---------- メインヘッダー ---------- */
.main-header {
    background: linear-gradient(135deg, #1a1a2e 0%, #16213e 50%, #0f3460 100%);
    color: #fff;
    padding: 1.5rem 2rem;
    border-radius: 16px;
    margin-bottom: 1.5rem;
    box-shadow: 0 8px 32px rgba(0,0,0,.15);
    position: relative;
    overflow: hidden;
}
.main-header::before {
    content: '';
    position: absolute;
    top: -50%; right: -50%;
    width: 100%; height: 200%;
    background: radial-gradient(circle, rgba(102,126,234,.12) 0%, transparent 70%);
}
.main-header h1 { margin:0; font-size:1.6rem; font-weight:700; position:relative; }
.main-header p  { margin:.4rem 0 0; font-size:.85rem; opacity:.8; position:relative; }

/* ---------- ステップヘッダー ---------- */
.step-header {
    background: linear-gradient(135deg, #f8f9ff 0%, #eef1ff 100%);
    border-left: 4px solid #667eea;
    padding: 1rem 1.5rem;
    border-radius: 0 12px 12px 0;
    margin-bottom: 1.2rem;
}
.step-header h2 { margin:0; font-size:1.15rem; font-weight:600; color:#1a1a2e; }
.step-header p  { margin:.3rem 0 0; font-size:.8rem; color:#666; }

/* ---------- カード ---------- */
.info-card {
    background: #fff;
    border: 1px solid #e8ecf1;
    border-radius: 12px;
    padding: 1.2rem 1.5rem;
    margin: .5rem 0;
    box-shadow: 0 2px 8px rgba(0,0,0,.04);
}

/* ---------- メトリクスカード ---------- */
.metric-card {
    background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
    color: #fff;
    padding: 1.1rem;
    border-radius: 12px;
    text-align: center;
    box-shadow: 0 4px 15px rgba(102,126,234,.3);
}
.metric-card .number { font-size:1.9rem; font-weight:700; line-height:1; }
.metric-card .label  { font-size:.78rem; opacity:.9; margin-top:.25rem; }
.mc-green  { background: linear-gradient(135deg, #11998e 0%, #38ef7d 100%);
              box-shadow: 0 4px 15px rgba(17,153,142,.3); }
.mc-orange { background: linear-gradient(135deg, #f5af19 0%, #f12711 100%);
              box-shadow: 0 4px 15px rgba(245,175,25,.3); }
.mc-pink   { background: linear-gradient(135deg, #f093fb 0%, #f5576c 100%);
              box-shadow: 0 4px 15px rgba(240,147,251,.3); }

/* ---------- サイドバー ---------- */
.progress-step {
    display: flex; align-items: center;
    padding: 9px 13px; margin: 3px 0;
    border-radius: 9px; font-size:.83rem;
}
.progress-done    { background:#e8f5e9; color:#2e7d32; border-left:3px solid #4caf50; }
.progress-pending { background:#f5f5f5; color:#9e9e9e; border-left:3px solid #e0e0e0; }

/* ---------- アップロードゾーン ---------- */
.upload-zone {
    border: 2px dashed #667eea;
    border-radius: 20px;
    padding: 2.5rem 2rem;
    text-align: center;
    background: linear-gradient(135deg, #f8f9ff 0%, #eef1ff 100%);
    margin: 1rem 0;
}
.upload-zone .icon { font-size:2.8rem; margin-bottom:.4rem; }
.upload-zone .text { color:#555; font-size:.92rem; }

/* ---------- カタログバッジ ---------- */
.catalog-badge {
    display: inline-block;
    background: linear-gradient(135deg, #e3f2fd 0%, #bbdefb 100%);
    color: #1565c0;
    padding: 3px 10px;
    border-radius: 20px;
    font-size: .75rem;
    font-weight: 500;
}

/* ---------- 区切り線 ---------- */
.gradient-divider {
    height: 3px;
    background: linear-gradient(90deg, #667eea 0%, #764ba2 50%, transparent 100%);
    border: none;
    margin: 1.8rem 0 1.3rem;
    border-radius: 2px;
}

/* ---------- タブ ---------- */
.stTabs [data-baseweb="tab-list"] {
    gap: 4px;
    background: linear-gradient(135deg, #f0f2f6 0%, #e8eaf0 100%);
    padding: 5px; border-radius: 14px;
}
.stTabs [data-baseweb="tab"] {
    border-radius: 10px; padding: 9px 18px; font-weight: 500; font-size:.88rem;
}
.stTabs [data-baseweb="tab"][aria-selected="true"] {
    background: #fff; box-shadow: 0 2px 8px rgba(0,0,0,.1);
}

/* ---------- ボタン ---------- */
div.stDownloadButton > button {
    background: linear-gradient(135deg, #11998e 0%, #38ef7d 100%) !important;
    border: none !important; border-radius: 10px !important;
    color: #fff !important; font-weight: 600 !important;
    box-shadow: 0 4px 15px rgba(17,153,142,.3) !important;
}

#MainMenu {visibility:hidden;}
footer   {visibility:hidden;}
</style>
""",
    unsafe_allow_html=True,
)

# =========================================================
# ヘルパー関数
# =========================================================

def extract_catalog_images(file_bytes):
    """xlsx 内の画像付きシートを自動検出し、シート名→画像bytesリストを返す。"""
    catalogs = {}
    try:
        with zipfile.ZipFile(io.BytesIO(file_bytes), "r") as zf:
            # 1. メディアファイル（画像）の読み込み
            media = {}
            for name in zf.namelist():
                if "/media/" in name and not name.endswith("/"):
                    media[posixpath.basename(name)] = zf.read(name)
            if not media:
                return catalogs

            ns_ss  = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
            ns_r   = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
            ns_rel = "http://schemas.openxmlformats.org/package/2006/relationships"

            # 2. workbook.xml → シート rId→名前
            wb_xml = ET.fromstring(zf.read("xl/workbook.xml"))
            rid_name = {}
            for el in wb_xml.iter(f"{{{ns_ss}}}sheet"):
                rid_name[el.get(f"{{{ns_r}}}id")] = el.get("name")

            # 3. workbook.xml.rels → rId→ファイルパス
            wb_rels = ET.fromstring(zf.read("xl/_rels/workbook.xml.rels"))
            rid_target = {}
            for rel in wb_rels.iter(f"{{{ns_rel}}}Relationship"):
                rid_target[rel.get("Id")] = rel.get("Target")

            # 4. シートごとに drawing→画像を辿る
            for rid, sheet_name in rid_name.items():
                target = rid_target.get(rid, "")
                if not target:
                    continue
                sheet_path = f"xl/{target}" if not target.startswith("xl/") else target
                sheet_dir  = posixpath.dirname(sheet_path)
                rels_path  = f"{sheet_dir}/_rels/{posixpath.basename(sheet_path)}.rels"
                if rels_path not in zf.namelist():
                    continue

                for rel in ET.fromstring(zf.read(rels_path)).iter(f"{{{ns_rel}}}Relationship"):
                    if "drawing" not in rel.get("Type", ""):
                        continue
                    dt = rel.get("Target")
                    dp = posixpath.normpath(posixpath.join(sheet_dir, dt)) if dt.startswith("..") else (
                        f"xl/{dt}" if not dt.startswith("xl/") else dt
                    )
                    dd  = posixpath.dirname(dp)
                    dr  = f"{dd}/_rels/{posixpath.basename(dp)}.rels"
                    if dr not in zf.namelist():
                        continue

                    imgs = []
                    for ir in ET.fromstring(zf.read(dr)).iter(f"{{{ns_rel}}}Relationship"):
                        if "image" in ir.get("Type", ""):
                            fn = posixpath.basename(ir.get("Target"))
                            if fn in media:
                                imgs.append(media[fn])
                    if imgs:
                        catalogs.setdefault(sheet_name, []).extend(imgs)
    except Exception:
        pass
    return catalogs


def load_data(file_bytes):
    """拾い出しシートを読み込み、列名を正規化したDataFrameを返す。"""
    df = pd.read_excel(io.BytesIO(file_bytes), sheet_name="拾い出し")

    rename = {}
    for col in df.columns:
        s = str(col).strip()
        if "名称" in s:
            rename[col] = "名称"
        elif "材料" in s and ("コード" in s or "ｺｰﾄﾞ" in s):
            rename[col] = "材料コード"
        elif "発注業者" in s or "発注先" in s:
            rename[col] = "発注先"
        elif s == "担当者":
            rename[col] = "担当者"
        elif "規格" in s:
            if "14.5" in s:
                rename[col] = "規格"
            elif "8" in s:
                rename[col] = "規格_1"
            elif "6.75" in s or "入数" in s:
                rename[col] = "規格_2"
        elif "階" in s:
            rename[col] = "階"
        elif s.startswith("発注") and "予定日" not in s:
            rename[col] = "発注"
        elif "単位" in s:
            rename[col] = "単位"
        elif "納品日" in s:
            rename[col] = "納品日"
        elif "納品場所" in s:
            rename[col] = "納品場所"
        elif "納品備考" in s:
            rename[col] = "納品備考"

    df = df.rename(columns=rename)

    for c in ["材料コード","名称","発注先","担当者","規格","規格_1","規格_2",
              "階","発注","単位","納品日","納品場所","納品備考"]:
        if c not in df.columns:
            df[c] = ""

    df = df.dropna(subset=["名称"])
    df = df[df["名称"].astype(str).str.strip() != ""]
    df.insert(0, "発注対象", False)
    return df


def generate_order_excel(order_df, contractor, person_name, order_date, file_bytes):
    """発注書・連絡書の2シートだけを持つExcelバイナリを生成して返す。"""
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes))

    ws = wb["発注書"] if "発注書" in wb.sheetnames else wb.create_sheet("発注書")

    def safe_set(row, col, value):
        cell = ws.cell(row=row, column=col)
        if not isinstance(cell, openpyxl.cell.cell.MergedCell):
            cell.value = value

    if ws.max_row >= 2:
        for r in range(2, ws.max_row + 1):
            for c in range(1, 17):
                safe_set(r, c, None)

    date_str = order_date.strftime("%Y/%m/%d") if isinstance(order_date, (date, datetime)) else str(order_date)

    for idx, row in enumerate(order_df.itertuples(), start=2):
        safe_set(idx, 1, getattr(row, "材料コード", None))
        safe_set(idx, 2, date_str)
        safe_set(idx, 5, contractor)
        safe_set(idx, 6, person_name if person_name else getattr(row, "担当者", None))
        safe_set(idx, 7, getattr(row, "名称", None))
        safe_set(idx, 8, getattr(row, "規格", None))
        safe_set(idx, 9, getattr(row, "規格_1", None))
        safe_set(idx, 10, getattr(row, "規格_2", None))

        def _clean(v):
            return "" if pd.isna(v) or str(v) == "NaN" else v

        safe_set(idx, 11, _clean(getattr(row, "階", None)))
        safe_set(idx, 12, _clean(getattr(row, "発注", None)))
        safe_set(idx, 13, _clean(getattr(row, "単位", None)))

        nd = getattr(row, "納品日", None)
        if pd.notna(nd) and str(nd) != "NaN":
            safe_set(idx, 14, nd.strftime("%Y/%m/%d") if isinstance(nd, datetime) else str(nd).split(" ")[0])
        else:
            safe_set(idx, 14, "")

        safe_set(idx, 15, _clean(getattr(row, "納品場所", None)))
        safe_set(idx, 16, _clean(getattr(row, "納品備考", None)))

    # 連絡書の□調整
    if "連絡書" in wb.sheetnames:
        ws_r = wb["連絡書"]
        for i in range(14):
            if i >= len(order_df):
                ws_r.cell(row=5 + i, column=1).value = None
        ws_r.print_area = "A1:H20"

    # 不要シート削除
    for s in list(wb.sheetnames):
        if s not in ("発注書", "連絡書"):
            wb.remove(wb[s])
    if wb.sheetnames:
        wb.active = 0
    for v in wb.views:
        v.activeTab = 0
        v.firstSheet = 0

    buf = io.BytesIO()
    wb.save(buf)
    wb.close()
    buf.seek(0)
    return buf.getvalue()


def convert_excel_to_pdf(excel_bytes, sheet_filter="連絡書"):
    """ExcelバイナリをPDFバイナリに変換して返す。"""
    import tempfile

    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        tmp.write(excel_bytes)
        tmp_path = tmp.name

    # 指定シート以外を非表示
    if sheet_filter:
        wb_t = openpyxl.load_workbook(tmp_path)
        for s in list(wb_t.sheetnames):
            if s != sheet_filter:
                wb_t[s].sheet_state = "hidden"
        wb_t.save(tmp_path)
        wb_t.close()

    pdf_path = tmp_path.replace(".xlsx", ".pdf")

    try:
        if platform.system() == "Windows":
            try:
                import win32com.client
            except ImportError:
                raise Exception("Windows環境で win32com が見つかりません。pywin32 をインストールしてください。")
            excel_app = None
            wb = None
            try:
                excel_app = win32com.client.DispatchEx("Excel.Application")
                excel_app.Visible = False
                excel_app.DisplayAlerts = False
                wb = excel_app.Workbooks.Open(os.path.abspath(tmp_path))
                wb.ExportAsFixedFormat(0, os.path.abspath(pdf_path))
            finally:
                if wb:
                    wb.Close(False)
                if excel_app:
                    excel_app.Quit()
        else:
            subprocess.run(
                ["libreoffice", "--headless", "--convert-to", "pdf",
                 tmp_path, "--outdir", os.path.dirname(tmp_path)],
                check=True, stdout=subprocess.PIPE, stderr=subprocess.PIPE,
            )
            lo_pdf = tmp_path.replace(".xlsx", ".pdf")
            if lo_pdf != pdf_path and os.path.exists(lo_pdf):
                os.rename(lo_pdf, pdf_path)

        if os.path.exists(pdf_path):
            with open(pdf_path, "rb") as f:
                return f.read()
        raise Exception("PDFが生成されませんでした。")
    finally:
        for p in (tmp_path, pdf_path):
            try:
                if os.path.exists(p):
                    os.remove(p)
            except OSError:
                pass


# =========================================================
# パスワード認証
# =========================================================
def check_password():
    if "password_correct" not in st.session_state:
        st.session_state["password_correct"] = False

    if not st.session_state["password_correct"]:
        st.markdown(
            '<div class="main-header">'
            "<h1>📋 自動発注・連絡書システム</h1>"
            "<p>ログインしてシステムを利用してください</p>"
            "</div>",
            unsafe_allow_html=True,
        )
        _, center, _ = st.columns([1, 2, 1])
        with center:
            st.markdown(
                '<div class="info-card" style="text-align:center;padding:2rem;">'
                '<div style="font-size:3rem;margin-bottom:.8rem;">🔐</div>'
                '<h3 style="color:#1a1a2e;margin:0;">ログイン</h3>'
                "</div>",
                unsafe_allow_html=True,
            )
            pwd = st.text_input("パスワード", type="password", placeholder="パスワードを入力…")
            if pwd:
                if pwd == st.secrets.get("app_password", ""):
                    st.session_state["password_correct"] = True
                    st.rerun()
                else:
                    st.error("😕 パスワードが正しくありません。")
        return False
    return True


if not check_password():
    st.stop()

# =========================================================
# メインヘッダー
# =========================================================
st.markdown(
    '<div class="main-header">'
    "<h1>📋 自動発注・連絡書システム</h1>"
    "<p>拾い出し表のアップロードから発注書・連絡書の作成、PDF変換までを一括管理</p>"
    "</div>",
    unsafe_allow_html=True,
)

# =========================================================
# セッション初期化
# =========================================================
_defaults = dict(
    file_bytes=None,
    raw_df=None,
    display_df=None,
    catalog_images={},
    supplier_catalog_map={},
    selected_contractor="--選択してください--",
    generated_excel_bytes=None,
    generated_excel_name="",
    person_name="",
    order_date=date.today(),
)
for k, v in _defaults.items():
    if k not in st.session_state:
        st.session_state[k] = v

# 既存ファイルの自動読込（初回のみ）
if st.session_state.file_bytes is None and os.path.exists("拾い出し表.xlsx"):
    try:
        with open("拾い出し表.xlsx", "rb") as f:
            fb = f.read()
        st.session_state.file_bytes = fb
        st.session_state.raw_df = load_data(fb)
        st.session_state.display_df = st.session_state.raw_df.copy()
        st.session_state.catalog_images = extract_catalog_images(fb)
        suppliers = st.session_state.raw_df["発注先"].dropna().unique()
        mapping = {}
        for sup in suppliers:
            for cat in st.session_state.catalog_images:
                if str(sup).strip() == cat.strip() or cat in str(sup):
                    mapping[str(sup)] = cat
        st.session_state.supplier_catalog_map = mapping
    except Exception:
        pass

# =========================================================
# サイドバー
# =========================================================
with st.sidebar:
    st.markdown("### 📊 ワークフロー進捗")
    st.markdown('<div class="gradient-divider"></div>', unsafe_allow_html=True)

    _raw = st.session_state.raw_df
    steps_info = [
        ("📤", "Step 1", "Excelアップロード",  st.session_state.file_bytes is not None),
        ("📖", "Step 2", "カタログ閲覧",       len(st.session_state.catalog_images) > 0),
        ("🔍", "Step 3", "検索・チェック",     _raw is not None and bool(_raw["発注対象"].any()) if _raw is not None else False),
        ("📝", "Step 4", "発注情報入力",       st.session_state.selected_contractor != "--選択してください--"),
        ("📥", "Step 5", "Excel出力",          st.session_state.generated_excel_bytes is not None),
        ("📄", "Step 6", "PDF変換",            False),
    ]
    for icon, step, label, done in steps_info:
        cls = "progress-done" if done else "progress-pending"
        chk = "✅" if done else "⬜"
        st.markdown(
            f'<div class="progress-step {cls}">{icon} <strong>{step}</strong>&nbsp; {label} &nbsp;{chk}</div>',
            unsafe_allow_html=True,
        )

    if _raw is not None:
        st.markdown("---")
        st.markdown("### 📈 データサマリー")
        c1, c2 = st.columns(2)
        c1.metric("総材料数", len(_raw))
        c1.metric("チェック済", int(_raw["発注対象"].sum()))
        c2.metric("業者数", _raw["発注先"].dropna().nunique())
        c2.metric("カタログ", len(st.session_state.catalog_images))

# =========================================================
# メインタブ
# =========================================================
tab1, tab2, tab3, tab4, tab5, tab6 = st.tabs([
    "📤 Step 1: アップロード",
    "📖 Step 2: カタログ",
    "🔍 Step 3: 検索・チェック",
    "📝 Step 4: 発注情報",
    "📥 Step 5: Excel出力",
    "📄 Step 6: PDF変換",
])

# =========================================================
# Step 1 — Excelアップロード
# =========================================================
with tab1:
    st.markdown(
        '<div class="step-header"><h2>📤 拾い出し表のアップロード</h2>'
        "<p>拾い出し表のExcelファイル（.xlsx）をアップロードしてください</p></div>",
        unsafe_allow_html=True,
    )
    st.markdown(
        '<div class="upload-zone">'
        '<div class="icon">📁</div>'
        '<div class="text">下のボタンから拾い出し表（.xlsx）を選択してください</div>'
        "</div>",
        unsafe_allow_html=True,
    )

    uploaded = st.file_uploader("拾い出し表Excelファイル", type=["xlsx"], key="main_uploader")

    if uploaded is not None:
        uid_key = "_last_main_id"
        if st.session_state.get(uid_key) != uploaded.file_id:
            try:
                fb = uploaded.getbuffer().tobytes()
                st.session_state.file_bytes = fb
                st.session_state.raw_df = load_data(fb)
                st.session_state.display_df = st.session_state.raw_df.copy()
                st.session_state.catalog_images = extract_catalog_images(fb)

                suppliers = st.session_state.raw_df["発注先"].dropna().unique()
                mapping = {}
                for sup in suppliers:
                    for cat in st.session_state.catalog_images:
                        if str(sup).strip() == cat.strip() or cat in str(sup):
                            mapping[str(sup)] = cat
                st.session_state.supplier_catalog_map = mapping

                st.session_state.generated_excel_bytes = None
                st.session_state.generated_excel_name = ""
                st.session_state[uid_key] = uploaded.file_id

                # ディスクにも保存（互換用）
                with open("拾い出し表.xlsx", "wb") as f:
                    f.write(fb)

                st.toast("✅ ファイルを読み込みました！", icon="🎉")
                st.rerun()
            except Exception as e:
                st.error(f"❌ 読み込み失敗: {e}")

    # サマリー
    if st.session_state.raw_df is not None:
        st.markdown('<div class="gradient-divider"></div>', unsafe_allow_html=True)
        st.success("✅ データが正常に読み込まれています")

        df = st.session_state.raw_df
        c1, c2, c3, c4 = st.columns(4)
        with c1:
            st.markdown(
                f'<div class="metric-card"><div class="number">{len(df)}</div>'
                '<div class="label">総材料数</div></div>', unsafe_allow_html=True)
        with c2:
            st.markdown(
                f'<div class="metric-card mc-green"><div class="number">{df["発注先"].dropna().nunique()}</div>'
                '<div class="label">発注先業者数</div></div>', unsafe_allow_html=True)
        with c3:
            st.markdown(
                f'<div class="metric-card mc-orange"><div class="number">{len(st.session_state.catalog_images)}</div>'
                '<div class="label">カタログシート数</div></div>', unsafe_allow_html=True)
        with c4:
            st.markdown(
                f'<div class="metric-card mc-pink"><div class="number">{len(st.session_state.supplier_catalog_map)}</div>'
                '<div class="label">カタログリンク数</div></div>', unsafe_allow_html=True)

        st.markdown("")
        if st.session_state.catalog_images:
            st.info(f"📖 カタログシート: {', '.join(st.session_state.catalog_images.keys())}")
        if st.session_state.supplier_catalog_map:
            links = [f"{k} → {v}" for k, v in st.session_state.supplier_catalog_map.items()]
            st.info(f"🔗 カタログリンク: {', '.join(links)}")

# =========================================================
# Step 2 — カタログ閲覧
# =========================================================
with tab2:
    st.markdown(
        '<div class="step-header"><h2>📖 カタログ閲覧</h2>'
        "<p>Excel内の画像付きシートをカタログとして表示します</p></div>",
        unsafe_allow_html=True,
    )

    if not st.session_state.catalog_images:
        st.warning("📂 カタログ画像がありません。Step 1 で拾い出し表をアップロードしてください。")
    else:
        cat_names = list(st.session_state.catalog_images.keys())

        default_idx = 0
        if "catalog_link_target" in st.session_state and st.session_state.catalog_link_target in cat_names:
            default_idx = cat_names.index(st.session_state.catalog_link_target)

        sel_cat = st.selectbox("📂 カタログシートを選択", cat_names, index=default_idx, key="cat_sel")

        if sel_cat:
            imgs = st.session_state.catalog_images[sel_cat]
            st.markdown(f"**{sel_cat}** — {len(imgs)} 枚の画像")
            st.markdown('<div class="gradient-divider"></div>', unsafe_allow_html=True)

            n_cols = st.slider("表示列数", 1, 4, 2, key="cat_cols")
            for i in range(0, len(imgs), n_cols):
                cols = st.columns(n_cols)
                for j, col in enumerate(cols):
                    idx = i + j
                    if idx < len(imgs):
                        with col:
                            st.image(imgs[idx], caption=f"{sel_cat} - 画像 {idx+1}",
                                     use_container_width=True)

# =========================================================
# Step 3 — 検索・チェック
# =========================================================
with tab3:
    st.markdown(
        '<div class="step-header"><h2>🔍 材料の検索・チェック</h2>'
        "<p>発注する材料にチェックを入れてください。検索で絞り込みも可能です。</p></div>",
        unsafe_allow_html=True,
    )

    if st.session_state.raw_df is None:
        st.warning("📂 データがありません。Step 1 で拾い出し表をアップロードしてください。")
    else:
        search_q = st.text_input("🔍 絞り込み検索", "", placeholder="業者名、材料名、材料コード、規格で検索…", key="search_q")

        raw = st.session_state.raw_df
        if search_q:
            mask = (
                raw["発注先"].astype(str).str.contains(search_q, case=False, na=False)
                | raw["名称"].astype(str).str.contains(search_q, case=False, na=False)
                | raw["材料コード"].astype(str).str.contains(search_q, case=False, na=False)
                | raw["規格"].astype(str).str.contains(search_q, case=False, na=False)
            )
            disp = raw[mask].copy()
        else:
            disp = raw.copy()
        st.session_state.display_df = disp

        mc1, mc2, mc3 = st.columns([2, 2, 6])
        mc1.metric("表示中", f"{len(disp)} 件")
        mc2.metric("チェック済（全体）", f"{int(raw['発注対象'].sum())} 件")
        if st.session_state.supplier_catalog_map:
            mc3.caption(f"📖 カタログあり: {', '.join(st.session_state.supplier_catalog_map.keys())}")

        def _on_edit(editor_key):
            es = st.session_state.get(editor_key, {})
            for pos, edits in es.get("edited_rows", {}).items():
                if "発注対象" in edits:
                    real_idx = st.session_state.display_df.index[int(pos)]
                    st.session_state.raw_df.at[real_idx, "発注対象"] = edits["発注対象"]

        vis_cols = ["発注対象", "材料コード", "名称", "規格", "規格_1", "発注", "単位", "発注先"]
        ekey = f"mat_editor_{search_q}"

        st.data_editor(
            disp,
            column_order=vis_cols,
            hide_index=True,
            use_container_width=True,
            disabled=["材料コード", "名称", "規格", "規格_1", "発注", "単位", "発注先"],
            key=ekey,
            on_change=_on_edit,
            kwargs={"editor_key": ekey},
        )

        # 関連カタログ — クイックビュー
        if st.session_state.supplier_catalog_map:
            st.markdown('<div class="gradient-divider"></div>', unsafe_allow_html=True)
            st.markdown("### 📖 関連カタログ — クイックビュー")

            vis_sups = disp["発注先"].dropna().unique()
            related = {str(s): st.session_state.supplier_catalog_map[str(s)]
                       for s in vis_sups if str(s) in st.session_state.supplier_catalog_map}

            if related:
                sel_sup = st.selectbox("業者を選択してカタログを表示", list(related.keys()), key="rel_cat_sel")
                if sel_sup:
                    cn = related[sel_sup]
                    cimgs = st.session_state.catalog_images.get(cn, [])
                    with st.expander(f"📖 {cn} のカタログ（{len(cimgs)} 枚）", expanded=True):
                        ec = st.columns(min(len(cimgs), 3)) if cimgs else []
                        for i, img in enumerate(cimgs):
                            with ec[i % len(ec)]:
                                st.image(img, caption=f"画像 {i+1}", use_container_width=True)
            else:
                st.info("表示中の材料に関連するカタログはありません。")

# =========================================================
# Step 4 — 発注情報入力
# =========================================================
with tab4:
    st.markdown(
        '<div class="step-header"><h2>📝 発注情報の入力</h2>'
        "<p>発注先、担当者名、日付を設定してください</p></div>",
        unsafe_allow_html=True,
    )

    if st.session_state.raw_df is None:
        st.warning("📂 データがありません。Step 1 で拾い出し表をアップロードしてください。")
    else:
        raw = st.session_state.raw_df
        checked = raw[raw["発注対象"] == True]

        if checked.empty:
            st.warning("⚠️ 発注対象の材料がチェックされていません。Step 3 でチェックしてください。")
        else:
            st.success(f"✅ {len(checked)} 件の材料がチェック済みです")

            c1, c2, c3 = st.columns(3)
            with c1:
                contractors = raw["発注先"].dropna().unique().tolist()
                sel = st.selectbox("🏢 発注先業者", ["--選択してください--"] + contractors, key="ctr_sel")
                st.session_state.selected_contractor = sel
            with c2:
                pn = st.text_input("👤 担当者名", value=st.session_state.person_name,
                                   placeholder="担当者の名前…", key="pn_inp")
                st.session_state.person_name = pn
            with c3:
                od = st.date_input("📅 発注日", value=st.session_state.order_date, key="dt_inp")
                st.session_state.order_date = od

            st.markdown('<div class="gradient-divider"></div>', unsafe_allow_html=True)

            if sel != "--選択してください--":
                odf = checked.copy()
                odf["発注先"] = sel
                odf = odf.sort_values("材料コード")

                st.markdown("### 📋 発注書プレビュー")
                st.markdown(f"**{sel}** 宛て ／ 担当者: **{pn or '(未設定)'}** ／ 日付: **{od.strftime('%Y/%m/%d')}**")
                st.dataframe(odf[["材料コード","名称","規格","規格_1","発注","単位"]],
                             hide_index=True, use_container_width=True)

                # カタログリンク
                ss = str(sel)
                if ss in st.session_state.supplier_catalog_map:
                    cn = st.session_state.supplier_catalog_map[ss]
                    st.markdown(f'<span class="catalog-badge">📖 カタログあり: {cn}</span>',
                                unsafe_allow_html=True)
                    if st.button(f"📖 {cn} のカタログを表示", key="cat_from_s4"):
                        st.session_state.catalog_link_target = cn
                        if cn in st.session_state.catalog_images:
                            with st.expander(f"📖 {cn} カタログ", expanded=True):
                                ci = st.session_state.catalog_images[cn]
                                ec = st.columns(min(len(ci), 3))
                                for i, img in enumerate(ci):
                                    with ec[i % len(ec)]:
                                        st.image(img, caption=f"画像 {i+1}", use_container_width=True)
            else:
                st.info("👆 発注先業者を選択してください")

# =========================================================
# Step 5 — Excel出力
# =========================================================
with tab5:
    st.markdown(
        '<div class="step-header"><h2>📥 発注書・連絡書Excel出力</h2>'
        "<p>設定した内容でExcelを生成してダウンロードします</p></div>",
        unsafe_allow_html=True,
    )

    if st.session_state.raw_df is None:
        st.warning("📂 Step 1 で拾い出し表をアップロードしてください。")
    elif st.session_state.selected_contractor == "--選択してください--":
        st.warning("⚠️ Step 4 で発注先を選択してください。")
    else:
        raw = st.session_state.raw_df
        checked = raw[raw["発注対象"] == True].copy()

        if checked.empty:
            st.warning("⚠️ 発注対象の材料がチェックされていません。")
        else:
            ctr = st.session_state.selected_contractor
            pn  = st.session_state.person_name
            od  = st.session_state.order_date
            checked["発注先"] = ctr
            checked = checked.sort_values("材料コード")

            st.markdown(
                f'<div class="info-card">'
                f"<strong>🏢 発注先:</strong> {ctr} &nbsp;|&nbsp;"
                f"<strong>👤 担当者:</strong> {pn or '(未設定)'} &nbsp;|&nbsp;"
                f"<strong>📅 日付:</strong> {od.strftime('%Y/%m/%d')} &nbsp;|&nbsp;"
                f"<strong>📦 材料数:</strong> {len(checked)} 件"
                "</div>",
                unsafe_allow_html=True,
            )
            st.markdown("")

            out_name = f"発注書_連絡書_{ctr}.xlsx"

            if st.button("📄 Excelを発行する", type="primary", use_container_width=True, key="gen_xl"):
                try:
                    xl = generate_order_excel(checked, ctr, pn, od, st.session_state.file_bytes)
                    st.session_state.generated_excel_bytes = xl
                    st.session_state.generated_excel_name = out_name
                    st.toast("✅ Excelを発行しました！", icon="🎉")
                    st.rerun()
                except Exception as e:
                    st.error(f"❌ Excel生成エラー: {e}")

            if st.session_state.generated_excel_bytes:
                st.markdown('<div class="gradient-divider"></div>', unsafe_allow_html=True)
                st.success(f"✅ 「{st.session_state.generated_excel_name}」を発行しました！")
                st.info("※ このファイルには『発注書』シートと『連絡書』シートが含まれています。")
                st.download_button(
                    label=f"📥 「{st.session_state.generated_excel_name}」をダウンロード",
                    data=st.session_state.generated_excel_bytes,
                    file_name=st.session_state.generated_excel_name,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                    key="dl_xl",
                )

# =========================================================
# Step 6 — PDF変換
# =========================================================
with tab6:
    st.markdown(
        '<div class="step-header"><h2>📄 ExcelからPDFへ変換</h2>'
        "<p>Excelファイルをアップロードし、名前を決めてPDFに変換します</p></div>",
        unsafe_allow_html=True,
    )

    pdf_source = None
    pdf_source_name = ""

    # ショートカット: Step 5 で生成済みの場合
    if st.session_state.generated_excel_bytes:
        use_gen = st.checkbox(
            f"✅ Step 5 で生成した「{st.session_state.generated_excel_name}」をそのまま使う",
            value=True,
            key="use_gen_chk",
        )
        if use_gen:
            pdf_source = st.session_state.generated_excel_bytes
            pdf_source_name = st.session_state.generated_excel_name

    if pdf_source is None:
        st.markdown(
            '<div class="upload-zone">'
            '<div class="icon">📄</div>'
            '<div class="text">PDF変換するExcelファイルを選択してください</div>'
            "</div>",
            unsafe_allow_html=True,
        )
        pdf_up = st.file_uploader("PDF変換用Excelファイル", type=["xlsx"], key="pdf_up")
        if pdf_up:
            pdf_source = pdf_up.getbuffer().tobytes()
            pdf_source_name = pdf_up.name

    if pdf_source:
        st.success(f"✅ 「{pdf_source_name}」を使用します")

        try:
            wb_chk = openpyxl.load_workbook(io.BytesIO(pdf_source))
            sheets = wb_chk.sheetnames
            wb_chk.close()
        except Exception as e:
            st.error(f"❌ ファイル読み込みエラー: {e}")
            st.stop()

        c1, c2 = st.columns(2)
        with c1:
            def_idx = sheets.index("連絡書") + 1 if "連絡書" in sheets else 0
            pdf_sheet = st.selectbox("📋 PDFに含めるシート", ["すべて"] + sheets,
                                     index=def_idx, key="pdf_sh_sel")
        with c2:
            base = os.path.splitext(pdf_source_name)[0]
            now_s = datetime.now().strftime("%Y%m%d_%H%M")
            pdf_name = st.text_input("📝 PDFファイル名", value=f"{base}_{now_s}.pdf", key="pdf_nm")

        if not pdf_name.endswith(".pdf"):
            pdf_name += ".pdf"

        st.markdown('<div class="gradient-divider"></div>', unsafe_allow_html=True)

        if st.button("🔄 PDFに変換する", type="primary", use_container_width=True, key="conv_pdf"):
            with st.spinner("ExcelからPDFへ変換中… ⏳"):
                try:
                    filt = None if pdf_sheet == "すべて" else pdf_sheet
                    pdf_bytes = convert_excel_to_pdf(pdf_source, filt)
                    st.session_state["pdf_result"] = pdf_bytes
                    st.session_state["pdf_result_name"] = pdf_name
                    st.toast("✅ PDF変換が完了しました！", icon="📄")
                    st.rerun()
                except Exception as e:
                    st.error(f"❌ PDF変換エラー: {e}")

        if st.session_state.get("pdf_result"):
            rn = st.session_state.get("pdf_result_name", "output.pdf")
            st.success(f"✅ 「{rn}」の変換が完了しました！")
            st.download_button(
                label=f"📥 「{rn}」をダウンロード",
                data=st.session_state["pdf_result"],
                file_name=rn,
                mime="application/pdf",
                use_container_width=True,
                key="dl_pdf",
            )
    elif not st.session_state.generated_excel_bytes:
        st.info("💡 まず Step 5 でExcelを発行するか、ここにExcelファイルをアップロードしてください。")
